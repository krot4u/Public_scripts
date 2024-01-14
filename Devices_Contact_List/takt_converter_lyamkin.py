import openpyxl
from openpyxl.styles import PatternFill

# pip install -r requirements.txt

# Переменные для имён файлов
file_new_contact = "DMRIds.dat"
file_takt_contact = "Takt_Contact_En.xls"
file_takt_contact_rus = "Takt_Contact_Ru.xls"


with open(file_new_contact, 'r') as f:
    lines2 = [line.split('\t') for line in [line.strip() for line in f]]
f.close()


book = openpyxl.Workbook()
book.remove(book.active)
sheet1 = book.create_sheet("Basic Information")
sheet2 = book.create_sheet("DMR Services_Contact")
sheet3 = book.create_sheet("HDC1200 Services_Contact")
sheet4 = book.create_sheet("2-Tone Services_Contact")
sheet5 = book.create_sheet("5-Tone Services_Contact")
sheet6 = book.create_sheet("5-Tone Services_Contact List")
sheet7 = book.create_sheet("Phone_List")


for row in lines2:
    sheet2.append(row)


for c in sheet2['A']:
    new_cell = c.offset(column=3)
    new_cell.value = c.value


sheet2.insert_rows(0)
sheet2["B1"].value = "P25"
sheet2["D1"].value = "94006"
sheet2.insert_rows(0)
sheet2["B1"].value = "FRN"
sheet2["D1"].value = "94005"
sheet2.insert_rows(0)
sheet2["B1"].value = "Попугай"
sheet2["D1"].value = "94004"
sheet2.insert_rows(0)
sheet2["B1"].value = "Шашлычная"
sheet2["D1"].value = "94003"
sheet2.insert_rows(0)
sheet2["B1"].value = "Кабак"
sheet2["D1"].value = "94002"
sheet2.insert_rows(0)
sheet2["B1"].value = "Курилка"
sheet2["D1"].value = "94001"
sheet2.insert_rows(0)
sheet2["B1"].value = "TG-5973"
sheet2["D1"].value = "5973"
sheet2.insert_rows(0)
sheet2["B1"].value = "QRA-Team"
sheet2["D1"].value = "9"
sheet2.insert_rows(0)
sheet2["A1"].value = "No."
sheet2["B1"].value = "Call Alias"
sheet2["C1"].value = "Call Type"
sheet2["D1"].value = "Call ID"
sheet1["A1"].value = "Radio Data Version"
sheet1["B1"].value = "92"
sheet1["A2"].value = "Model Type"
sheet1["B2"].value = "Mobile"
sheet1["A3"].value = "Sale Region"
sheet1["B3"].value = "5"
sheet1["A4"].value = "Dial Rules"
sheet1["B4"].value = "None"


sheet1.merge_cells('A5:B5')
megre_cell = sheet1['A5']
megre_cell.value = "QRA-Team forever"
megre_cell.fill = PatternFill('solid', fgColor="AAFF00")


for col in range(2, sheet2.max_row+1):
    sheet2['A' + str(col)].value = col-1
    sheet2['C' + str(col)].value = 'Private Call'
sheet2["C2"].value = "Group Call"
u_id = u_alias = sheet2['B']
id1list=[]
for x in range(len(u_alias)):
    if u_id[x].value==None:
        continue
    r = (u_alias[x].value, u_alias[x].value)
    if r in id1list:
        u_alias[x].value = u_alias[x].value + str(1)
        id1list.append(r)
    else:
        id1list.append(r)
book.save(file_takt_contact)


book = openpyxl.Workbook()
book.remove(book.active)
sheet1 = book.create_sheet("Базовая информация")
sheet2 = book.create_sheet("Сервисы DMR_Контакты")
sheet3 = book.create_sheet("Сервисы HDC1200_Контакты")
sheet4 = book.create_sheet("2-тоновые службы_Контакты")
sheet5 = book.create_sheet("5-Тоновые службы_Контакты")
sheet7 = book.create_sheet("Телефония_Список")


for row in lines2:
    sheet2.append(row)


for c in sheet2['A']:
    new_cell = c.offset(column=3)
    new_cell.value = c.value


sheet2.insert_rows(0)
sheet2["B1"].value = "P25"
sheet2["D1"].value = "94006"
sheet2.insert_rows(0)
sheet2["B1"].value = "FRN"
sheet2["D1"].value = "94005"
sheet2.insert_rows(0)
sheet2["B1"].value = "Попугай"
sheet2["D1"].value = "94004"
sheet2.insert_rows(0)
sheet2["B1"].value = "Шашлычная"
sheet2["D1"].value = "94003"
sheet2.insert_rows(0)
sheet2["B1"].value = "Кабак"
sheet2["D1"].value = "94002"
sheet2.insert_rows(0)
sheet2["B1"].value = "Курилка"
sheet2["D1"].value = "94001"
sheet2.insert_rows(0)
sheet2["B1"].value = "TG-5973"
sheet2["D1"].value = "5973"
sheet2.insert_rows(0)
sheet2["B1"].value = "QRA-Team"
sheet2["D1"].value = "9"
sheet2.insert_rows(0)
sheet2["A1"].value = "№"
sheet2["B1"].value = "Имя"
sheet2["C1"].value = "Тип вызова"
sheet2["D1"].value = "ID"
sheet1["A1"].value = "Версия сигнатуры"
sheet1["B1"].value = "92"
sheet1["A2"].value = "Тип модели"
sheet1["B2"].value = "Автомобильная"
sheet1["A3"].value = "Sale Region"
sheet1["B3"].value = "5"
sheet1["A4"].value = "Правила набора"
sheet1["B4"].value = "Нет"


sheet1.merge_cells('A5:B5')
megre_cell = sheet1['A5']
megre_cell.value = "QRA-Team forever"
megre_cell.fill = PatternFill('solid', fgColor="AAFF00")


for col in range(2, sheet2.max_row+1):
    sheet2['A' + str(col)].value = col-1
    sheet2['C' + str(col)].value = 'Индивидуальный вызов'
sheet2["C2"].value = "Групповой вызов"


u_id = u_alias = sheet2['B']
id1list=[]
for x in range(len(u_alias)):
    if u_id[x].value==None:
        continue
    r = (u_alias[x].value, u_alias[x].value)
    if r in id1list:
        u_alias[x].value = u_alias[x].value + str(1)
        id1list.append(r)
    else:
        id1list.append(r)


book.save(file_takt_contact_rus)
